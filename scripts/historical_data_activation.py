#!/usr/bin/env python3
"""Build a historical paike/finance activation package from local Excel files."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median

import openpyxl


ROOT = Path("/Users/chengzhihao/Desktop/排课系统")
OUT_DIR = Path("work/historical-data-activation")
YEAR = 2026
SKIP_SHEET_RE = re.compile(r"工资表|工资|产值|日常|消费|课程大纲|续报")
EMPTY_COURSE_RE = re.compile(r"^(可排|空|无|无课|休|休息|请假|停课|日期|时间|星期|周[一二三四五六日天])$")
SUPPORT_TASK_RE = re.compile(r"磨课|出门测|答疑|直播|托管|学习出")
TEACHER_ALIAS_SUFFIX_RE = re.compile(r"老师$")
ROOM_RE = re.compile(r"([一二三四五六七八九十\d]+楼\s*[一二三四五六七八九十\d]+号(?:教室)?|[一二三四五六七八九十\d]+楼[一二三四五六七八九十\d]+号(?:教室)?|[A-Za-z]?\d{2,4}教室?)")
TIME_RANGE_RE = re.compile(r"(\d{1,2})\s*[:：;]\s*(\d{1,2})\s*[-~至到]\s*(\d{1,2})\s*[:：;]\s*(\d{1,2})")
TIME_RE = re.compile(r"(\d{1,2})\s*[:：;]\s*(\d{1,2})")


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\u00a0", " ").strip()


def teacher_key(value: str) -> str:
    return TEACHER_ALIAS_SUFFIX_RE.sub("", re.sub(r"\s+", "", text(value))).strip()


def infer_teacher_from_file(path: Path) -> str:
    match = re.search(r"[（(]([^（）()]{2,8}?)(?:老师)?[）)]", path.name)
    if match:
        return teacher_key(match.group(1))
    stem = re.sub(r"\.(xlsx|xls)$", "", path.name, flags=re.I)
    stem = re.sub(r"2026.*$", "", stem)
    return teacher_key(stem)


def infer_period_from_name(value: str) -> str:
    source = text(value)
    match = re.search(r"20?26[.\-年 ]+([1-9]|1[0-2])\s*月?", source)
    if match:
        return f"{YEAR}-{int(match.group(1)):02d}"
    match = re.search(r"([1-9]|1[0-2])\s*月", source)
    if match:
        return f"{YEAR}-{int(match.group(1)):02d}"
    return ""


def month_from_text(value: str) -> int:
    period = infer_period_from_name(value)
    return int(period[-2:]) if period else 0


def excel_serial_to_date(value) -> str:
    try:
        number = float(value)
    except Exception:
        return ""
    if number < 1 or number > 60000:
        return ""
    base = dt.datetime(1899, 12, 30)
    return (base + dt.timedelta(days=int(number))).strftime("%Y-%m-%d")


def valid_month_day(month: int, day: int) -> bool:
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return False
    try:
        dt.date(YEAR, month, day)
        return True
    except ValueError:
        return False


def date_key(value, context_month: int = 0, allow_day_only: bool = False) -> str:
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = excel_serial_to_date(value)
        if serial:
            return serial
    raw = text(value)
    if not raw:
        return ""
    if re.search(r"\d{1,2}\s*[:：]\s*\d{1,2}", raw):
        return ""
    normalized = raw.replace("/", "-").replace(".", "-").replace("年", "-").replace("月", "-").replace("日", "")
    match = re.search(r"(20\d{2})-(\d{1,2})-(\d{1,2})", normalized)
    if match:
        month, day = int(match.group(2)), int(match.group(3))
        return f"{match.group(1)}-{month:02d}-{day:02d}" if valid_month_day(month, day) else ""
    match = re.search(r"(?:^|[^\d])(\d{1,2})-(\d{1,2})(?:$|[^\d])", normalized)
    if match:
        month, day = int(match.group(1)), int(match.group(2))
        return f"{YEAR}-{month:02d}-{day:02d}" if valid_month_day(month, day) else ""
    if allow_day_only and context_month:
        match = re.fullmatch(r"\s*(\d{1,2})(?:号|日)?\s*", raw)
        if match:
            day = int(match.group(1))
            return f"{YEAR}-{context_month:02d}-{day:02d}" if valid_month_day(context_month, day) else ""
    return ""


def normalize_time(value: str) -> str:
    match = TIME_RE.search(text(value))
    if not match:
        return ""
    return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}"


def time_range(value: str) -> tuple[str, str]:
    raw = text(value)
    match = TIME_RANGE_RE.search(raw)
    if match:
        start_hour, start_minute = int(match.group(1)), int(match.group(2))
        end_hour, end_minute = int(match.group(3)), int(match.group(4))
        if 1 <= start_hour <= 7 and 2 <= end_hour <= 9:
            start_hour += 12
            end_hour += 12
        return f"{start_hour:02d}:{start_minute:02d}", f"{end_hour:02d}:{end_minute:02d}"
    first = normalize_time(raw)
    return first, ""


def time_minutes(value: str) -> int:
    normalized = normalize_time(value)
    if not normalized:
        return 9999
    hour, minute = normalized.split(":")
    return int(hour) * 60 + int(minute)


def add_hours(value: str, hours: int) -> str:
    normalized = normalize_time(value)
    if not normalized:
        return ""
    hour, minute = normalized.split(":")
    return f"{(int(hour) + hours) % 24:02d}:{int(minute):02d}"


def align_cell_time_to_row(cell_start: str, cell_end: str, row_start: str, row_end: str) -> tuple[str, str]:
    if not cell_start or not row_start:
        return cell_start, cell_end
    cell_minutes = time_minutes(cell_start)
    row_minutes = time_minutes(row_start)
    if cell_minutes == 9999 or row_minutes == 9999:
        return cell_start, cell_end
    if row_minutes >= 12 * 60 and cell_minutes < 8 * 60:
        return add_hours(cell_start, 12), add_hours(cell_end, 12) if cell_end else cell_end
    if row_minutes >= 18 * 60 and cell_minutes < 12 * 60:
        return add_hours(cell_start, 12), add_hours(cell_end, 12) if cell_end else cell_end
    return cell_start, cell_end


def lesson_hours(start: str, end: str) -> float:
    left, right = time_minutes(start), time_minutes(end)
    if left != 9999 and right != 9999 and right > left:
        return round((right - left) / 60, 2)
    return 1.5


def row_left_text(row, date_start_col: int) -> str:
    return " ".join(text(cell) for cell in row[: max(date_start_col, 1)] if text(cell))


def has_time_rows_below(matrix, row_index: int, date_start_col: int) -> bool:
    count = 0
    for row in matrix[row_index + 1 : row_index + 13]:
        start, _ = time_range(row_left_text(row, date_start_col))
        if start:
            count += 1
    return count >= 1


def find_date_headers(matrix, sheet_name: str) -> list[dict]:
    headers = []
    context_month = month_from_text(sheet_name)
    for row_index, row in enumerate(matrix):
        row_text = " ".join(text(cell) for cell in row if text(cell))
        row_month = month_from_text(row_text)
        if row_month:
            context_month = row_month
        dates = []
        for col_index, value in enumerate(row):
            parsed = date_key(value, context_month=context_month, allow_day_only=True)
            if parsed:
                dates.append({"columnIndex": col_index, "date": parsed, "raw": text(value)})
        if len(dates) >= 5 and has_time_rows_below(matrix, row_index, min(item["columnIndex"] for item in dates)):
            previous = headers[-1] if headers else None
            if previous and row_index - previous["rowIndex"] <= 1:
                if len(dates) > len(previous["dates"]):
                    previous["dates"] = dates
                continue
            headers.append({"rowIndex": row_index, "dates": dates})
    return headers


def is_course_cell(value: str) -> bool:
    raw = text(value)
    if not raw or EMPTY_COURSE_RE.match(raw):
        return False
    stripped = TIME_RANGE_RE.sub("", raw).strip()
    if not stripped or EMPTY_COURSE_RE.match(stripped):
        return False
    return bool(re.search(r"[\u4e00-\u9fa5A-Za-z0-9]", raw))


def infer_room(*values: str) -> str:
    for value in values:
        match = ROOM_RE.search(text(value))
        if match:
            return re.sub(r"\s+", "", match.group(1))
    return ""


def clean_course_name(raw: str) -> str:
    result = TIME_RANGE_RE.sub("", text(raw))
    result = ROOM_RE.sub("", result)
    result = re.sub(r"\s+", " ", result).strip()
    return result or text(raw)


def lesson_category(raw: str) -> str:
    source = text(raw)
    if re.search(r"科学", source):
        return "科学"
    if re.search(r"班课|小班|刷题|集训|群课", source):
        return "小班/班课"
    if re.search(r"一对一|1对1", source):
        return "一对一"
    if SUPPORT_TASK_RE.search(source):
        return "教研/测评/辅助任务"
    return "待识别"


def activation_track(raw: str) -> str:
    source = text(raw)
    if SUPPORT_TASK_RE.search(source):
        return "support_task"
    if re.search(r"免费|试听", source):
        return "trial_or_marketing"
    return "formal_lesson"


def matrix_from_sheet(ws) -> list[list]:
    rows = []
    max_col = ws.max_column or 1
    for row in ws.iter_rows(values_only=True):
        values = list(row[:max_col])
        if any(text(cell) for cell in values):
            rows.append(values)
        else:
            rows.append(values)
    return rows


def parse_schedule_file(path: Path) -> tuple[list[dict], list[dict]]:
    teacher = infer_teacher_from_file(path)
    entries = []
    diagnostics = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        sheet_name = ws.title
        if SKIP_SHEET_RE.search(f"{path.name} {sheet_name}"):
            diagnostics.append({"file": path.name, "sheet": sheet_name, "status": "skipped", "reason": "非排课sheet"})
            continue
        matrix = matrix_from_sheet(ws)
        headers = find_date_headers(matrix, sheet_name)
        sheet_count = 0
        for block_index, header in enumerate(headers):
            next_header = headers[block_index + 1] if block_index + 1 < len(headers) else None
            date_start_col = min(item["columnIndex"] for item in header["dates"])
            block_rows = matrix[header["rowIndex"] + 1 : next_header["rowIndex"] if next_header else len(matrix)]
            for offset, row in enumerate(block_rows):
                absolute_row = header["rowIndex"] + 1 + offset
                start, end = time_range(row_left_text(row, date_start_col))
                date_cells = [text(row[item["columnIndex"]] if item["columnIndex"] < len(row) else "") for item in header["dates"]]
                has_course_cell = any(is_course_cell(value) for value in date_cells)
                if not start and not has_course_cell and sheet_count:
                    break
                if not start and not has_course_cell:
                    continue
                for date_item in header["dates"]:
                    col = date_item["columnIndex"]
                    raw = text(row[col] if col < len(row) else "")
                    if not is_course_cell(raw):
                        continue
                    cell_start, cell_end = time_range(raw)
                    cell_start, cell_end = align_cell_time_to_row(cell_start, cell_end, start, end)
                    final_start = cell_start or start
                    final_end = cell_end or end
                    if not final_start:
                        continue
                    header_text = " ".join(
                        text(matrix[idx][col] if 0 <= idx < len(matrix) and col < len(matrix[idx]) else "")
                        for idx in [header["rowIndex"] - 1, header["rowIndex"]]
                    )
                    room = infer_room(raw, header_text)
                    course_name = clean_course_name(raw)
                    entry = {
                        "id": re.sub(r"[^\w\u4e00-\u9fa5-]", "", f"hist-{path.name}-{sheet_name}-{date_item['date']}-{final_start}-{absolute_row}-{col}"),
                        "teacherName": teacher,
                        "teacher": teacher,
                        "className": course_name,
                        "studentName": course_name,
                        "date": date_item["date"],
                        "courseDate": date_item["date"],
                        "period": date_item["date"][:7],
                        "startTime": final_start,
                        "endTime": final_end,
                        "time": f"{final_start}-{final_end}" if final_end else final_start,
                        "classroomName": room,
                        "roomName": room,
                        "lessonTypeRaw": raw,
                        "rawText": raw,
                        "lessonCategory": lesson_category(raw),
                        "activationTrack": activation_track(raw),
                        "isBillableCandidate": activation_track(raw) == "formal_lesson",
                        "estimatedHours": lesson_hours(final_start, final_end),
                        "scheduleStatus": "scheduled",
                        "confirmationStatus": "confirmed",
                        "importedByExcel": True,
                        "sourceFileName": path.name,
                        "sourceSheet": sheet_name,
                        "sourceCell": f"R{absolute_row + 1}C{col + 1}",
                        "confidence": 0.92 if room else 0.84,
                    }
                    entries.append(entry)
                    sheet_count += 1
        diagnostics.append({
            "file": path.name,
            "sheet": sheet_name,
            "status": "parsed" if sheet_count else "empty",
            "headers": len(headers),
            "entries": sheet_count,
        })
    return entries, diagnostics


def known_teacher_names(schedule_files: list[Path]) -> list[str]:
    return sorted({infer_teacher_from_file(path) for path in schedule_files if infer_teacher_from_file(path)}, key=len, reverse=True)


def parse_finance_file(path: Path, teachers: list[str]) -> tuple[list[dict], list[dict]]:
    rows = []
    diagnostics = []
    default_period = infer_period_from_name(path.name)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    teacher_patterns = [(name, re.compile(re.escape(name))) for name in teachers]
    for ws in wb.worksheets:
        sheet_name = ws.title
        sheet_period = infer_period_from_name(f"{path.name} {sheet_name}") or default_period
        count = 0
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [text(cell) for cell in row]
            row_text = " ".join(cell for cell in cells if cell)
            if not row_text:
                continue
            matched_teacher = ""
            for name, pattern in teacher_patterns:
                if pattern.search(row_text):
                    matched_teacher = name
                    break
            if not matched_teacher:
                continue
            numeric_values = []
            for cell in row:
                if isinstance(cell, (int, float)) and not isinstance(cell, bool) and math.isfinite(float(cell)):
                    numeric_values.append(float(cell))
                else:
                    raw = text(cell).replace(",", "")
                    if re.fullmatch(r"-?\d+(?:\.\d+)?", raw):
                        numeric_values.append(float(raw))
            rows.append({
                "file": path.name,
                "sheet": sheet_name,
                "rowIndex": row_index,
                "period": sheet_period,
                "teacherName": matched_teacher,
                "rawText": row_text,
                "numericValues": numeric_values,
                "numericSum": round(sum(numeric_values), 2),
                "maxNumber": max(numeric_values) if numeric_values else 0,
                "confidence": 0.72 if sheet_period else 0.58,
            })
            count += 1
        diagnostics.append({"file": path.name, "sheet": sheet_name, "period": sheet_period, "financeRows": count})
    return rows, diagnostics


FINANCE_TEACHER_ALIAS = {
    "何": "何建军",
    "李舒老师": "李舒",
    "何老师": "何建军",
    "曹": "曹德顺",
    "曹老师": "曹德顺",
    "潘": "潘云贵",
    "潘老师": "潘云贵",
    "吴女": "吴水琴",
    "吴（女）": "吴水琴",
    "吴(女)": "吴水琴",
}


def numeric(value, default: float = 0.0) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value)):
        return float(value)
    raw = text(value).replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", raw):
        return float(raw)
    return default


def normalize_finance_teacher_name(value: str, known_teachers: list[str]) -> str:
    raw = text(value)
    if raw in FINANCE_TEACHER_ALIAS:
        return FINANCE_TEACHER_ALIAS[raw]
    key = teacher_key(raw).replace("(", "（").replace(")", "）")
    compact = re.sub(r"[（）()]", "", key)
    if key in FINANCE_TEACHER_ALIAS:
        return FINANCE_TEACHER_ALIAS[key]
    if compact in FINANCE_TEACHER_ALIAS:
        return FINANCE_TEACHER_ALIAS[compact]
    for teacher in known_teachers:
        if key == teacher_key(teacher) or key in teacher_key(teacher) or teacher_key(teacher) in key:
            return teacher
    return key


def period_month(period: str) -> int:
    match = re.search(r"-(\d{2})$", text(period))
    return int(match.group(1)) if match else 0


def lesson_dates_from_cells(cells: list, fallback_period: str) -> list[dict]:
    dates = []
    context_month = period_month(fallback_period)
    for cell in cells:
        raw = text(cell)
        if not raw:
            continue
        parsed = date_key(raw, context_month=context_month, allow_day_only=False)
        if parsed:
            dates.append({"date": parsed, "raw": raw})
    return dates


def find_header_indexes(row: list) -> dict:
    result = {}
    for index, cell in enumerate(row):
        value = text(cell)
        if value and value not in result:
            result[value] = index
    return result


def parse_finance_structured(path: Path, known_teachers: list[str]) -> tuple[dict, list[dict]]:
    summary_rows = []
    detail_rows = []
    diagnostics = []
    default_period = infer_period_from_name(path.name)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        sheet_name = ws.title
        sheet_period = infer_period_from_name(f"{path.name} {sheet_name}") or default_period
        if "小课产值" in sheet_name:
            headers = []
            header_row_index = 0
            for header_row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = list(row)
                if any(text(cell) == "姓名" for cell in values) and any(text(cell) == "课时总收入" for cell in values):
                    headers = values
                    break
            header_index = find_header_indexes(headers)
            if not header_index:
                diagnostics.append({"file": path.name, "sheet": sheet_name, "status": "summary_header_missing"})
                continue
            count = 0
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_index <= header_row_index:
                    continue
                values = list(row)
                name = text(values[header_index.get("姓名", -1)] if header_index.get("姓名", -1) < len(values) else "")
                if name in {"姓名"}:
                    continue
                if name == "合计":
                    break
                if not name:
                    if count:
                        break
                    continue
                if not re.search(r"[\u4e00-\u9fa5]", name):
                    continue
                teacher = normalize_finance_teacher_name(name, known_teachers)
                summary_rows.append({
                    "file": path.name,
                    "sheet": sheet_name,
                    "rowIndex": row_index,
                    "period": sheet_period,
                    "teacherName": teacher,
                    "lessonIncome": numeric(values[header_index.get("课时总收入", -1)] if header_index.get("课时总收入", -1) < len(values) else None),
                    "baseSalary": numeric(values[header_index.get("基础工资", -1)] if header_index.get("基础工资", -1) < len(values) else None),
                    "socialSecurity": numeric(values[header_index.get("社保", -1)] if header_index.get("社保", -1) < len(values) else None),
                    "classCommission": numeric(values[header_index.get("课时提成", -1)] if header_index.get("课时提成", -1) < len(values) else None),
                    "makeupCommission": numeric(values[header_index.get("补课提成", -1)] if header_index.get("补课提成", -1) < len(values) else None),
                    "qaCommission": numeric(values[header_index.get("答疑提成", -1)] if header_index.get("答疑提成", -1) < len(values) else None),
                    "subsidy": numeric(values[header_index.get("补贴", -1)] if header_index.get("补贴", -1) < len(values) else None),
                    "remaining": numeric(values[header_index.get("课时剩余", -1)] if header_index.get("课时剩余", -1) < len(values) else None),
                    "rawText": " ".join(text(cell) for cell in values if text(cell)),
                    "confidence": 0.9 if sheet_period else 0.7,
                })
                count += 1
            diagnostics.append({"file": path.name, "sheet": sheet_name, "status": "summary_parsed", "rows": count})
            continue

        rows = list(ws.iter_rows(values_only=True))
        header_row_index = -1
        header_index = {}
        for index, row in enumerate(rows):
            values = list(row)
            if any(text(cell) == "姓名" for cell in values) and any(text(cell) == "提成" for cell in values) and any(text(cell) == "计数" for cell in values):
                header_row_index = index
                header_index = find_header_indexes(values)
                break
        if header_row_index < 0:
            diagnostics.append({"file": path.name, "sheet": sheet_name, "status": "detail_skipped"})
            continue
        teacher = normalize_finance_teacher_name(sheet_name, known_teachers)
        name_col = header_index.get("姓名", 1)
        fee_col = header_index.get("课时费", 2)
        commission_col = header_index.get("提成", 3)
        count_col = header_index.get("计数", 12)
        total_col = header_index.get("总计课时", 13)
        date_start_col = min(index for label, index in header_index.items() if label in {"上课日期", "第1次"} or label.startswith("第")) if header_index else 4
        date_end_col = count_col
        count = 0
        for row_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            values = list(row)
            student = text(values[name_col] if name_col < len(values) else "")
            if not student or student in {"姓名", "合计"}:
                continue
            if not re.search(r"[\u4e00-\u9fa5]", student):
                continue
            date_cells = values[date_start_col:date_end_col]
            lesson_dates = lesson_dates_from_cells(date_cells, sheet_period)
            lesson_count_cell = numeric(values[count_col] if count_col < len(values) else None)
            lesson_count_by_dates = len(lesson_dates)
            lesson_count = int(lesson_count_cell or lesson_count_by_dates or 0)
            fee = numeric(values[fee_col] if fee_col < len(values) else None)
            commission = numeric(values[commission_col] if commission_col < len(values) else None)
            total_commission = numeric(values[total_col] if total_col < len(values) else None)
            computed_commission_total = round(commission * lesson_count, 2) if commission and lesson_count else 0
            if not lesson_count and not fee and not commission and not lesson_dates:
                continue
            detail_rows.append({
                "file": path.name,
                "sheet": sheet_name,
                "rowIndex": row_index,
                "period": sheet_period,
                "teacherName": teacher,
                "studentName": student,
                "lessonFee": fee,
                "commissionPerLesson": commission,
                "lessonCount": lesson_count,
                "lessonCountByDates": lesson_count_by_dates,
                "lessonDates": lesson_dates,
                "lessonIncome": round(fee * lesson_count, 2) if fee and lesson_count else 0,
                "commissionTotal": computed_commission_total,
                "cachedTotalCell": total_commission,
                "rawText": " ".join(text(cell) for cell in values if text(cell))[:500],
                "confidence": 0.93 if lesson_dates else 0.78,
            })
            count += 1
        diagnostics.append({"file": path.name, "sheet": sheet_name, "status": "detail_parsed", "teacherName": teacher, "rows": count})
    return {"summaryRows": summary_rows, "detailRows": detail_rows}, diagnostics


STUDENT_TOKEN_BLACKLIST = {
    "补初", "补四", "补五", "补六", "春季", "暑假", "寒假", "数学", "科学", "小班", "班课", "一对", "免费", "试听",
    "年级", "全册", "奥数", "出门", "刷题", "托管", "课程", "时间", "加课", "第次", "初一", "初二", "初三",
}


def student_tokens_from_course(raw: str) -> list[str]:
    source = clean_course_name(raw)
    source = re.split(r"[（(]", source, maxsplit=1)[0]
    source = TIME_RANGE_RE.sub("", source)
    parts = re.split(r"[+＋、，,/\s]+", source)
    tokens = []
    for part in parts:
        part = re.sub(r"[^\u4e00-\u9fa5]", "", part)
        if not (2 <= len(part) <= 4):
            continue
        if part in STUDENT_TOKEN_BLACKLIST or any(bad in part for bad in STUDENT_TOKEN_BLACKLIST):
            continue
        tokens.append(part)
    return list(dict.fromkeys(tokens))


def build_schedule_student_index(entries: list[dict]) -> dict:
    index = {}
    no_student_count = 0
    participant_count = 0
    for entry in entries:
        if entry.get("activationTrack") != "formal_lesson":
            continue
        tokens = student_tokens_from_course(entry.get("className") or entry.get("rawText") or "")
        if not tokens:
            no_student_count += 1
            continue
        for student in tokens:
            key = (entry.get("teacherName") or entry.get("teacher") or "", entry.get("period") or "", student)
            item = index.setdefault(key, {
                "teacherName": key[0],
                "period": key[1],
                "studentName": key[2],
                "scheduleCount": 0,
                "dates": set(),
                "samples": [],
            })
            item["scheduleCount"] += 1
            participant_count += 1
            if entry.get("date"):
                item["dates"].add(entry["date"])
            if len(item["samples"]) < 3:
                item["samples"].append({
                    "date": entry.get("date"),
                    "time": entry.get("time"),
                    "className": entry.get("className"),
                    "source": f"{entry.get('sourceFileName')} {entry.get('sourceSheet')} {entry.get('sourceCell')}",
                })
    for item in index.values():
        item["dates"] = sorted(item["dates"])
    return {"index": index, "noStudentLessonCount": no_student_count, "participantCount": participant_count}


def aggregate_finance_structured(structured: dict) -> dict:
    summary_by_teacher_month = {}
    detail_by_teacher_month = {}
    detail_by_student = {}
    for row in structured.get("summaryRows", []):
        key = (row["teacherName"], row["period"])
        summary_by_teacher_month[key] = row
    for row in structured.get("detailRows", []):
        key = (row["teacherName"], row["period"])
        item = detail_by_teacher_month.setdefault(key, {
            "teacherName": key[0],
            "period": key[1],
            "studentRows": 0,
            "lessonCount": 0,
            "lessonIncome": 0.0,
            "commissionTotal": 0.0,
            "datedLessonCount": 0,
        })
        item["studentRows"] += 1
        item["lessonCount"] += int(row.get("lessonCount") or 0)
        item["lessonIncome"] += float(row.get("lessonIncome") or 0)
        item["commissionTotal"] += float(row.get("commissionTotal") or 0)
        item["datedLessonCount"] += int(row.get("lessonCountByDates") or 0)
        student_key = (row["teacherName"], row["period"], row["studentName"])
        detail_by_student[student_key] = row
    for item in detail_by_teacher_month.values():
        item["lessonIncome"] = round(item["lessonIncome"], 2)
        item["commissionTotal"] = round(item["commissionTotal"], 2)
    return {
        "summaryByTeacherMonth": summary_by_teacher_month,
        "detailByTeacherMonth": detail_by_teacher_month,
        "detailByStudent": detail_by_student,
    }


def build_reconciliation(entries: list[dict], finance_structured: dict) -> dict:
    schedule_index_info = build_schedule_student_index(entries)
    schedule_index = schedule_index_info["index"]
    finance_agg = aggregate_finance_structured(finance_structured)
    detail_by_student = finance_agg["detailByStudent"]
    detail_by_teacher_month = finance_agg["detailByTeacherMonth"]
    summary_by_teacher_month = finance_agg["summaryByTeacherMonth"]
    student_matches = []
    auto_decisions = Counter()
    for key, finance_row in detail_by_student.items():
        schedule_row = schedule_index.get(key)
        finance_dates = sorted({item["date"] for item in finance_row.get("lessonDates", []) if item.get("date")})
        schedule_dates = schedule_row.get("dates", []) if schedule_row else []
        matched_dates = sorted(set(finance_dates) & set(schedule_dates))
        finance_count = int(finance_row.get("lessonCount") or 0)
        schedule_count = int(schedule_row.get("scheduleCount") or 0) if schedule_row else 0
        if finance_count == schedule_count and (not finance_dates or len(matched_dates) == len(finance_dates)):
            status = "matched"
            decision = "工资明细和排课表可互证，系统直接通过。"
        elif schedule_count == 0 and finance_count > 0:
            status = "finance_only"
            decision = "工资表有明确学生和上课日期，系统以工资表作为历史结算凭证；排课侧不强行补伪课程。"
        elif finance_count == 0 and schedule_count > 0:
            status = "schedule_only"
            decision = "排课表有课程但工资表未形成结算，系统保留排课并标记为未结算候选。"
        elif abs(finance_count - schedule_count) <= 1:
            status = "minor_diff"
            decision = "次数只差 1，系统按工资表结算、按排课表保留课程，进入低风险自动解释。"
        else:
            status = "major_diff"
            decision = "次数差异较大，系统不阻塞并网；结算采用工资表，排课采用课表，并进入AI复核池。"
        auto_decisions[status] += 1
        student_matches.append({
            "teacherName": key[0],
            "period": key[1],
            "studentName": key[2],
            "financeCount": finance_count,
            "scheduleCount": schedule_count,
            "matchedDateCount": len(matched_dates),
            "financeDates": finance_dates,
            "scheduleDates": schedule_dates,
            "status": status,
            "systemDecision": decision,
            "financeCommissionTotal": finance_row.get("commissionTotal", 0),
            "financeLessonIncome": finance_row.get("lessonIncome", 0),
            "scheduleSamples": schedule_row.get("samples", []) if schedule_row else [],
        })
    for key, schedule_row in schedule_index.items():
        if key in detail_by_student:
            continue
        teacher, period, student = key
        if period not in {"2026-05", "2026-06"}:
            continue
        student_matches.append({
            "teacherName": teacher,
            "period": period,
            "studentName": student,
            "financeCount": 0,
            "scheduleCount": schedule_row.get("scheduleCount", 0),
            "matchedDateCount": 0,
            "financeDates": [],
            "scheduleDates": schedule_row.get("dates", []),
            "status": "schedule_only",
            "systemDecision": "排课表有课程但工资表无结算记录，系统保留排课，不自动生成应付工资。",
            "financeCommissionTotal": 0,
            "financeLessonIncome": 0,
            "scheduleSamples": schedule_row.get("samples", []),
        })
        auto_decisions["schedule_only"] += 1

    monthly = []
    teacher_month_keys = sorted(set(detail_by_teacher_month.keys()) | set(summary_by_teacher_month.keys()))
    for key in teacher_month_keys:
        detail = detail_by_teacher_month.get(key, {"lessonCount": 0, "lessonIncome": 0, "commissionTotal": 0, "studentRows": 0, "datedLessonCount": 0})
        summary = summary_by_teacher_month.get(key, {})
        income_diff = round(float(summary.get("lessonIncome") or 0) - float(detail.get("lessonIncome") or 0), 2)
        commission_diff = round(float(summary.get("classCommission") or 0) - float(detail.get("commissionTotal") or 0), 2)
        if abs(income_diff) <= 1 and abs(commission_diff) <= 1:
            status = "summary_detail_matched"
            decision = "汇总表与老师明细表一致，系统直接通过。"
        elif not summary.get("lessonIncome") and detail.get("lessonIncome"):
            status = "summary_not_calculated"
            decision = "汇总表暂未计算或未保存公式结果，系统采用老师明细表作为财务依据。"
        else:
            status = "summary_detail_diff"
            decision = "汇总表与明细表有差异，系统保留两套凭证，结算判断优先采用明细表。"
        monthly.append({
            "teacherName": key[0],
            "period": key[1],
            "detailStudentRows": detail.get("studentRows", 0),
            "detailLessonCount": detail.get("lessonCount", 0),
            "detailLessonIncome": detail.get("lessonIncome", 0),
            "detailCommissionTotal": detail.get("commissionTotal", 0),
            "summaryLessonIncome": summary.get("lessonIncome", 0),
            "summaryClassCommission": summary.get("classCommission", 0),
            "incomeDiff": income_diff,
            "commissionDiff": commission_diff,
            "status": status,
            "systemDecision": decision,
        })
    return {
        "scheduleParticipantCount": schedule_index_info["participantCount"],
        "scheduleNoStudentLessonCount": schedule_index_info["noStudentLessonCount"],
        "studentMatches": sorted(student_matches, key=lambda row: (row["status"], row["period"], row["teacherName"], row["studentName"])),
        "monthly": monthly,
        "autoDecisionSummary": dict(auto_decisions),
    }


def detect_conflicts(entries: list[dict]) -> list[dict]:
    conflicts = []
    buckets = defaultdict(list)
    for entry in entries:
        key = (entry.get("teacherName"), entry.get("date"), entry.get("startTime"), entry.get("endTime"))
        buckets[key].append(entry)
    for key, group in buckets.items():
        if len(group) > 1:
            conflicts.append({
                "type": "teacher_time_duplicate",
                "severity": "high",
                "teacherName": key[0],
                "date": key[1],
                "time": f"{key[2]}-{key[3]}",
                "count": len(group),
                "samples": [item.get("className") for item in group[:5]],
            })
    room_buckets = defaultdict(list)
    for entry in entries:
        room = entry.get("roomName") or entry.get("classroomName")
        if not room:
            continue
        key = (room, entry.get("date"), entry.get("startTime"), entry.get("endTime"))
        room_buckets[key].append(entry)
    for key, group in room_buckets.items():
        if len(group) > 1:
            conflicts.append({
                "type": "room_time_duplicate",
                "severity": "medium",
                "roomName": key[0],
                "date": key[1],
                "time": f"{key[2]}-{key[3]}",
                "count": len(group),
                "samples": [f"{item.get('teacherName')} {item.get('className')}" for item in group[:5]],
            })
    return conflicts


def aggregate_schedule(entries: list[dict]) -> dict:
    result = {}
    for entry in entries:
        key = (entry.get("teacherName", ""), entry.get("period", ""))
        item = result.setdefault(key, {
            "teacherName": key[0],
            "period": key[1],
            "lessonCount": 0,
            "estimatedHours": 0.0,
            "missingRoomCount": 0,
            "categories": Counter(),
        })
        item["lessonCount"] += 1
        item["estimatedHours"] += float(entry.get("estimatedHours") or 0)
        if not entry.get("roomName") and not entry.get("classroomName"):
            item["missingRoomCount"] += 1
        item["categories"][entry.get("lessonCategory") or "待识别"] += 1
    for item in result.values():
        item["estimatedHours"] = round(item["estimatedHours"], 2)
        item["categories"] = dict(item["categories"])
    return result


def aggregate_finance(rows: list[dict]) -> dict:
    result = {}
    for row in rows:
        key = (row.get("teacherName", ""), row.get("period", ""))
        item = result.setdefault(key, {
            "teacherName": key[0],
            "period": key[1],
            "rowCount": 0,
            "numericSums": [],
            "maxNumbers": [],
            "rawSamples": [],
        })
        item["rowCount"] += 1
        item["numericSums"].append(float(row.get("numericSum") or 0))
        item["maxNumbers"].append(float(row.get("maxNumber") or 0))
        if len(item["rawSamples"]) < 3:
            item["rawSamples"].append(row.get("rawText", "")[:180])
    for item in result.values():
        item["numericSumTotal"] = round(sum(item["numericSums"]), 2)
        item["maxNumberMedian"] = round(median(item["maxNumbers"]), 2) if item["maxNumbers"] else 0
        del item["numericSums"]
        del item["maxNumbers"]
    return result


def build_risks(schedule_summary: dict, finance_summary: dict, conflicts: list[dict], entries: list[dict]) -> list[dict]:
    risks = []
    for conflict in conflicts[:200]:
        risks.append({
            "severity": conflict["severity"],
            "area": "排课冲突",
            "title": f"{conflict['type']} {conflict.get('date')} {conflict.get('time')}",
            "detail": json.dumps(conflict, ensure_ascii=False),
            "action": "系统先保留并网；同格多课按低置信度处理，进入AI复核池。",
        })
    for key, item in schedule_summary.items():
        if item["lessonCount"] >= 5 and item["missingRoomCount"] / max(1, item["lessonCount"]) > 0.6:
            risks.append({
                "severity": "low",
                "area": "教室",
                "title": f"{item['teacherName']} {item['period']} 未写教室较多",
                "detail": f"{item['missingRoomCount']}/{item['lessonCount']} 节未写教室。",
                "action": "系统自动保留课程，不阻塞并网；只降低教室利用率分析置信度。",
            })
        if key not in finance_summary and item["period"] in {"2026-05", "2026-06"}:
            risks.append({
                "severity": "medium",
                "area": "财务匹配",
                "title": f"{item['teacherName']} {item['period']} 有排课但未匹配工资表",
                "detail": f"排课 {item['lessonCount']} 节，估算 {item['estimatedHours']} 小时。",
                "action": "排课照常进入系统；财务侧不自动生成工资，只作为未结算候选。",
            })
    for key, item in finance_summary.items():
        if key not in schedule_summary:
            risks.append({
                "severity": "medium",
                "area": "财务匹配",
                "title": f"{item['teacherName']} {item['period']} 有工资记录但未匹配排课",
                "detail": f"工资表命中 {item['rowCount']} 行。",
                "action": "系统采用工资表作为历史财务凭证，排课侧不强行补伪课程。",
            })
    for entry in entries:
        if not entry.get("date") or not entry.get("startTime") or not entry.get("teacherName"):
            risks.append({
                "severity": "high",
                "area": "基础字段",
                "title": "课程缺关键字段",
                "detail": json.dumps(entry, ensure_ascii=False)[:500],
                "action": "这类数据不进入正式课表，系统自动转入历史异常池。",
            })
    severity_order = {"high": 0, "medium": 1, "low": 2}
    return sorted(risks, key=lambda row: (severity_order.get(row["severity"], 9), row["area"], row["title"]))


def write_json(path: Path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    headers = sorted({key for row in rows for key in row.keys() if key != "numericValues"})
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value for key, value in row.items() if key in headers})


def scan_files(root: Path) -> dict:
    all_files = sorted(path for path in root.rglob("*") if path.is_file() and not path.name.startswith("."))
    inventory = []
    schedule_files = []
    finance_files = []
    old_files = []
    for path in all_files:
        relative = path.relative_to(root)
        kind = "other"
        is_archive = any("旧数据存档" in part or "编程不用读" in part for part in relative.parts)
        if is_archive:
            old_files.append(path)
            kind = "old_archive"
        elif re.search(r"工资|分红|课时费", path.name):
            finance_files.append(path)
            kind = "finance"
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            schedule_files.append(path)
            kind = "schedule"
        inventory.append({
            "path": str(path),
            "relativePath": str(relative),
            "fileName": path.name,
            "kind": kind,
            "size": path.stat().st_size,
        })
    return {
        "inventory": inventory,
        "scheduleFiles": schedule_files,
        "financeFiles": finance_files,
        "oldFiles": old_files,
    }


def make_report(summary: dict, schedule_summary: dict, finance_summary: dict, risks: list[dict], output_paths: dict, reconciliation: dict | None = None) -> str:
    high = sum(1 for risk in risks if risk["severity"] == "high")
    medium = sum(1 for risk in risks if risk["severity"] == "medium")
    low = sum(1 for risk in risks if risk["severity"] == "low")
    top_schedule = sorted(schedule_summary.values(), key=lambda row: row["lessonCount"], reverse=True)[:12]
    lines = [
        "# 历史数据盘活第一阶段报告",
        "",
        "## 总体结论",
        f"- 扫描文件：{summary['fileCount']} 个",
        f"- 最新老师课表：{summary['scheduleFileCount']} 个",
        f"- 工资/分红表：{summary['financeFileCount']} 个",
        f"- 旧存档文件：{summary['oldFileCount']} 个（本轮不覆盖最新表，只作为后续补充）",
        f"- 识别课表单元：{summary['scheduleEntryCount']} 条",
        f"- 正式课程候选：{summary['formalLessonCount']} 节",
        f"- 教研/测评/辅助任务：{summary['supportTaskCount']} 条",
        f"- 试听/免费/市场课：{summary['trialOrMarketingCount']} 条",
        f"- 涉及老师：{summary['teacherCount']} 位",
        f"- 涉及月份：{', '.join(summary['periods']) or '-'}",
        f"- 命中财务行：{summary['financeRowCount']} 行",
        f"- 风险汇总：高 {high} / 中 {medium} / 低 {low}",
        "",
        "## 执行原则",
        "- 不要求老师逐条核对。",
        "- 排课表和工资表先自动互证，能解释的差异由系统直接处理。",
        "- 金额结算优先采用工资明细表；排课展示优先采用老师课表。",
        "- 无法确定的内容不阻塞系统，只降置信度并进入AI复核池。",
        "- 老师不参与繁琐核对，系统只把最终少量决策点留给程老师。",
        "",
        "## 对标成熟教培系统后纳入的原则",
        "- 业财融合：排课、课消、工资、财务不要分散成孤岛，所有数据必须能按老师/月度/学生回溯。",
        "- 入口前置：报名和续费完成后，应自动生成排课、收费、财务和后续服务线索，减少二次录入。",
        "- 自动排课校验：同一老师、同一教室、同一时间的冲突由系统自动检测和降置信度。",
        "- 月结锁定：历史月份进入月结状态后不随便改原始记录，后续变化走修正、冲销、补录或版本记录。",
        "- 风险标签：学生、家长、课程、老师都可以沉淀风险标签，但标签必须服务于续费、排课和服务决策。",
        "- AI减负：AI用于批量生成反馈、对账解释、经营分析和复盘建议，不把核对压力转移给老师。",
        "",
    ]
    if reconciliation:
        auto = reconciliation.get("autoDecisionSummary") or {}
        lines.extend([
            "## 排课-工资自动对账",
            f"- 工资明细行：{summary.get('financeDetailRowCount', 0)} 行",
            f"- 工资汇总行：{summary.get('financeSummaryRowCount', 0)} 行",
            f"- 排课学生课次识别：{reconciliation.get('scheduleParticipantCount', 0)} 次",
            f"- 可互证学生记录：{auto.get('matched', 0)} 条",
            f"- 工资表有、排课表未精确匹配：{auto.get('finance_only', 0)} 条，系统按工资表保留结算凭证",
            f"- 排课表有、工资表未结算：{auto.get('schedule_only', 0)} 条，系统保留课程但不生成应付工资",
            f"- 小差异自动解释：{auto.get('minor_diff', 0)} 条",
            f"- 大差异进入AI复核池：{auto.get('major_diff', 0)} 条",
            "",
            "## 月度财务口径",
        ])
        for row in (reconciliation.get("monthly") or [])[:20]:
            lines.append(
                f"- {row['teacherName']} {row['period']}：明细 {row['detailLessonCount']} 次，"
                f"明细提成 {row['detailCommissionTotal']}；汇总提成 {row['summaryClassCommission']}；{row['systemDecision']}"
            )
        lines.append("")
    lines.append("## 正式课程候选最多的老师/月度")
    for row in top_schedule:
        lines.append(f"- {row['teacherName']} {row['period']}：{row['lessonCount']} 节，约 {row['estimatedHours']} 小时，未写教室 {row['missingRoomCount']} 节")
    lines.extend(["", "## 最高优先级问题（前 30 条）"])
    for risk in risks[:30]:
        lines.append(f"- [{risk['severity']}] {risk['area']}｜{risk['title']}｜{risk['action']}")
    lines.extend([
        "",
        "## 输出文件",
        f"- 文件清单：{output_paths['inventory']}",
        f"- 统一排课 JSON：{output_paths['schedule_json']}",
        f"- 统一排课 CSV：{output_paths['schedule_csv']}",
        f"- 财务识别 JSON：{output_paths['finance_json']}",
        f"- 工资明细 JSON：{output_paths['finance_detail_json']}",
        f"- 自动对账 JSON：{output_paths['reconciliation_json']}",
        f"- 风险清单 JSON：{output_paths['risks_json']}",
        "",
        "## 下一步",
        "1. 不做老师逐条核对，系统按自动对账结果先盘活历史数据。",
        "2. 正式课程写入排课系统；工资明细写入历史财务候选库。",
        "3. 课程展示按课表，工资结算按工资明细，差异进入AI复核池。",
        "4. 网站里展示“系统已处理结论”，不把繁琐异常推给老师。",
    ])
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=str(ROOT))
    parser.add_argument("--out", default=str(OUT_DIR))
    args = parser.parse_args()
    root = Path(args.root)
    out_dir = Path(args.out)
    scanned = scan_files(root)
    schedule_files = scanned["scheduleFiles"]
    finance_files = scanned["financeFiles"]
    teachers = known_teacher_names(schedule_files)

    schedule_entries = []
    schedule_diagnostics = []
    for path in schedule_files:
        entries, diagnostics = parse_schedule_file(path)
        schedule_entries.extend(entries)
        schedule_diagnostics.extend(diagnostics)

    finance_rows = []
    finance_diagnostics = []
    finance_structured = {"summaryRows": [], "detailRows": []}
    finance_structured_diagnostics = []
    for path in finance_files:
        rows, diagnostics = parse_finance_file(path, teachers)
        finance_rows.extend(rows)
        finance_diagnostics.extend(diagnostics)
        structured, structured_diagnostics = parse_finance_structured(path, teachers)
        finance_structured["summaryRows"].extend(structured.get("summaryRows", []))
        finance_structured["detailRows"].extend(structured.get("detailRows", []))
        finance_structured_diagnostics.extend(structured_diagnostics)

    formal_entries = [entry for entry in schedule_entries if entry.get("activationTrack") == "formal_lesson"]
    schedule_summary = aggregate_schedule(formal_entries)
    finance_summary = aggregate_finance(finance_rows)
    reconciliation = build_reconciliation(schedule_entries, finance_structured)
    conflicts = detect_conflicts(formal_entries)
    risks = build_risks(schedule_summary, finance_summary, conflicts, formal_entries)

    periods = sorted({entry["period"] for entry in schedule_entries if entry.get("period")})
    teacher_names = sorted({entry["teacherName"] for entry in schedule_entries if entry.get("teacherName")})
    summary = {
        "fileCount": len(scanned["inventory"]),
        "scheduleFileCount": len(schedule_files),
        "financeFileCount": len(finance_files),
        "oldFileCount": len(scanned["oldFiles"]),
        "scheduleEntryCount": len(schedule_entries),
        "formalLessonCount": len(formal_entries),
        "supportTaskCount": sum(1 for entry in schedule_entries if entry.get("activationTrack") == "support_task"),
        "trialOrMarketingCount": sum(1 for entry in schedule_entries if entry.get("activationTrack") == "trial_or_marketing"),
        "financeRowCount": len(finance_rows),
        "financeSummaryRowCount": len(finance_structured["summaryRows"]),
        "financeDetailRowCount": len(finance_structured["detailRows"]),
        "reconciliationAutoDecisionSummary": reconciliation.get("autoDecisionSummary", {}),
        "reconciliationStudentMatchCount": len(reconciliation.get("studentMatches", [])),
        "teacherCount": len(teacher_names),
        "teachers": teacher_names,
        "periods": periods,
        "riskCount": len(risks),
        "conflictCount": len(conflicts),
    }

    paths = {
        "inventory": out_dir / "file_inventory.json",
        "schedule_json": out_dir / "schedule_entries.json",
        "schedule_csv": out_dir / "schedule_entries.csv",
        "finance_json": out_dir / "finance_rows.json",
        "finance_csv": out_dir / "finance_rows.csv",
        "finance_summary_json": out_dir / "finance_summary_rows.json",
        "finance_summary_csv": out_dir / "finance_summary_rows.csv",
        "finance_detail_json": out_dir / "finance_detail_rows.json",
        "finance_detail_csv": out_dir / "finance_detail_rows.csv",
        "reconciliation_json": out_dir / "reconciliation_report.json",
        "risks_json": out_dir / "risk_list.json",
        "summary_json": out_dir / "summary.json",
        "report": out_dir / "activation_report.md",
        "diagnostics": out_dir / "diagnostics.json",
    }
    write_json(paths["inventory"], scanned["inventory"])
    write_json(paths["schedule_json"], schedule_entries)
    write_csv(paths["schedule_csv"], schedule_entries)
    write_json(paths["finance_json"], finance_rows)
    write_csv(paths["finance_csv"], finance_rows)
    write_json(paths["finance_summary_json"], finance_structured["summaryRows"])
    write_csv(paths["finance_summary_csv"], finance_structured["summaryRows"])
    write_json(paths["finance_detail_json"], finance_structured["detailRows"])
    write_csv(paths["finance_detail_csv"], finance_structured["detailRows"])
    write_json(paths["reconciliation_json"], reconciliation)
    write_json(paths["risks_json"], risks)
    write_json(paths["summary_json"], summary)
    write_json(paths["diagnostics"], {
        "scheduleDiagnostics": schedule_diagnostics,
        "financeDiagnostics": finance_diagnostics,
        "financeStructuredDiagnostics": finance_structured_diagnostics,
        "scheduleSummary": {f"{k[0]}|{k[1]}": v for k, v in schedule_summary.items()},
        "financeSummary": {f"{k[0]}|{k[1]}": v for k, v in finance_summary.items()},
        "conflicts": conflicts,
    })
    report = make_report(summary, schedule_summary, finance_summary, risks, {key: str(value) for key, value in paths.items()}, reconciliation)
    paths["report"].write_text(report, encoding="utf-8")
    print(json.dumps({
        "ok": True,
        "summary": summary,
        "report": str(paths["report"]),
        "outDir": str(out_dir),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
